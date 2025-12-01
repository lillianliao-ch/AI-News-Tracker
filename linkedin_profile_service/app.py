import os
import re
import requests
from typing import List, Optional, Dict, Any
from datetime import datetime

from fastapi import FastAPI, HTTPException
from pydantic import BaseModel
from dotenv import load_dotenv
from openpyxl import Workbook
from typing import cast

try:
    # 可选导入，只有使用浏览器抓取模式时才需要
    from playwright.sync_api import sync_playwright
except Exception:  # pragma: no cover
    sync_playwright = None  # 延迟到运行时检查并报错


# 加载本地环境变量（可选）
load_dotenv()


app = FastAPI(title="LinkedIn Profile Parser (ContactOut)", version="0.1.0")


class ParseRequest(BaseModel):
    url: str
    provider: Optional[str] = None  # 可选：contactout（默认）或 serpapi


class TextParseRequest(BaseModel):
    text: str
    name: Optional[str] = None
    headline: Optional[str] = None
    location: Optional[str] = None


class BrowserParseRequest(BaseModel):
    url: str
    li_at: str
    jsessionid: Optional[str] = None
    headless: Optional[bool] = True


def is_linkedin_profile_url(url: str) -> bool:
    pattern = r"^https://(www\.)?linkedin\.com/in/[^/?#]+/?$"
    return bool(re.match(pattern, url))


def fetch_profile_from_contactout(profile_url: str, api_token: str) -> Dict[str, Any]:
    endpoint = "https://api.contactout.com/v1/linkedin/enrich"
    headers = {
        "authorization": "basic",
        "token": api_token,
    }
    params = {"profile": profile_url}

    try:
        resp = requests.get(endpoint, headers=headers, params=params, timeout=30)
    except requests.RequestException as exc:
        raise HTTPException(status_code=502, detail=f"上游网络错误: {exc}")

    if resp.status_code == 404:
        raise HTTPException(status_code=404, detail="未找到公开资料或链接不可访问")
    if resp.status_code == 401 or resp.status_code == 400:
        raise HTTPException(status_code=401, detail="鉴权失败或请求不合法，请检查 CONTACTOUT_API_TOKEN 与输入链接")
    if resp.status_code == 429:
        raise HTTPException(status_code=429, detail="达到速率限制，请稍后重试")
    if not resp.ok:
        raise HTTPException(status_code=502, detail=f"上游错误: HTTP {resp.status_code}")

    data = resp.json()
    # 检测 ContactOut 示例数据，避免误用
    sample_fullname = (data.get("profile") or {}).get("full_name")
    message_text = data.get("message", "")
    example_url = ((data.get("profile") or {}).get("url") or "")
    if (sample_fullname == "Example Person") or ("sample response" in message_text.lower()) or ("example-person" in example_url):
        raise HTTPException(status_code=402, detail="ContactOut 返回示例数据（沙盒/受限）。请更换正式 token 或切换其他数据提供方")
    return data


# SerpAPI 提供方（可选）
def fetch_profile_from_serpapi(profile_url: str, api_key: str) -> Dict[str, Any]:
    endpoint = "https://serpapi.com/search.json"
    params = {
        "engine": "linkedin_profile",
        "profile_url": profile_url,
        "api_key": api_key,
    }
    try:
        resp = requests.get(endpoint, params=params, timeout=45)
    except requests.RequestException as exc:
        raise HTTPException(status_code=502, detail=f"上游网络错误: {exc}")
    if resp.status_code == 401:
        raise HTTPException(status_code=401, detail="SerpAPI 鉴权失败，请检查 SERPAPI_KEY")
    if resp.status_code == 429:
        raise HTTPException(status_code=429, detail="SerpAPI 速率限制，请稍后重试")
    if not resp.ok:
        raise HTTPException(status_code=502, detail=f"SerpAPI 上游错误: HTTP {resp.status_code}")
    return resp.json()


def project_profile_from_serpapi(payload: Dict[str, Any]) -> Dict[str, Any]:
    # 尽量兼容不同字段名
    name = payload.get("full_name") or payload.get("name")
    headline = payload.get("headline") or payload.get("title")
    location = payload.get("location") or payload.get("country")

    experiences_src = payload.get("experiences") or payload.get("experience") or []
    experiences: List[Dict[str, Any]] = []
    for item in experiences_src or []:
        if not isinstance(item, dict):
            continue
        experiences.append({
            "company": item.get("company") or item.get("company_name"),
            "title": item.get("title") or item.get("position"),
            "start_date": item.get("start_date") or item.get("date_from") or item.get("starts_at"),
            "end_date": item.get("end_date") or item.get("date_to") or item.get("ends_at"),
            "is_current": item.get("is_current"),
            "location": item.get("location") or item.get("locality"),
            "summary": item.get("description") or item.get("summary"),
        })

    edus_src = payload.get("education") or payload.get("educations") or []
    educations: List[Dict[str, Any]] = []
    for edu in edus_src or []:
        if not isinstance(edu, dict):
            continue
        educations.append({
            "school": edu.get("school") or edu.get("school_name"),
            "degree": edu.get("degree") or edu.get("degree_name"),
            "field_of_study": edu.get("field_of_study") or edu.get("major"),
            "start_year": edu.get("start_year") or edu.get("start_date") or edu.get("date_from"),
            "end_year": edu.get("end_year") or edu.get("end_date") or edu.get("date_to"),
            "description": edu.get("description"),
        })

    return {
        "name": name,
        "headline": headline,
        "location": location,
        "experiences": experiences,
        "educations": educations,
    }


def resolve_provider(provider: Optional[str]) -> str:
    default_provider = os.getenv("DATA_PROVIDER", "contactout").lower()
    if provider:
        return provider.lower()
    return default_provider


def project_profile(payload: Dict[str, Any]) -> Dict[str, Any]:
    profile = payload.get("profile") or payload

    # 经验映射
    experiences = []
    for item in profile.get("experience", []) or []:
        if not isinstance(item, dict):
            continue
        experiences.append({
            "company": item.get("company_name"),
            "title": item.get("title"),
            "start_date": item.get("start_date") or _to_year_month(item),
            "end_date": item.get("end_date"),
            "is_current": item.get("is_current"),
            "location": item.get("locality"),
            "summary": item.get("summary"),
            "company_linkedin": item.get("linkedin_url"),
        })

    # 教育映射
    educations = []
    for edu in profile.get("education", []) or []:
        if not isinstance(edu, dict):
            continue
        educations.append({
            "school": edu.get("school_name"),
            "degree": edu.get("degree"),
            "field_of_study": edu.get("field_of_study"),
            "start_year": edu.get("start_date_year"),
            "end_year": edu.get("end_date_year"),
            "description": edu.get("description"),
        })

    result = {
        "name": profile.get("full_name"),
        "headline": profile.get("headline"),
        "location": profile.get("location") or profile.get("country"),
        "experiences": experiences,
        "educations": educations,
    }
    return result


def _to_year_month(item: Dict[str, Any]) -> Optional[str]:
    y = item.get("start_date_year")
    m = item.get("start_date_month")
    if y and m:
        return f"{int(y):04d}-{int(m):02d}"
    if y:
        return f"{int(y):04d}"
    return None


@app.post("/parse")
def parse_linkedin_profile(req: ParseRequest) -> Dict[str, Any]:
    if not is_linkedin_profile_url(req.url):
        raise HTTPException(status_code=400, detail="请输入合法的 LinkedIn 个人主页链接（形如 https://www.linkedin.com/in/<username>/ ）")

    provider = resolve_provider(req.provider)
    if provider == "serpapi":
        serp_key = os.getenv("SERPAPI_KEY")
        if not serp_key:
            raise HTTPException(status_code=500, detail="缺少 SERPAPI_KEY 环境变量")
        data = fetch_profile_from_serpapi(req.url, serp_key)
        return project_profile_from_serpapi(data)
    else:
        api_token = os.getenv("CONTACTOUT_API_TOKEN")
        if not api_token:
            raise HTTPException(status_code=500, detail="缺少 CONTACTOUT_API_TOKEN 环境变量")
        data = fetch_profile_from_contactout(req.url, api_token)
        return project_profile(data)


@app.get("/")
def root():
    return {"message": "OK", "service": "linkedin_profile_service", "version": "0.1.0"}


# ========== 增强功能：可读化文本 + Excel 导出 ==========

def _end_year_month(item: Dict[str, Any]) -> Optional[str]:
    y = item.get("end_date_year")
    m = item.get("end_date_month")
    if y and m:
        return f"{int(y):04d}-{int(m):02d}"
    if y:
        return f"{int(y):04d}"
    return None


def _format_period(item: Dict[str, Any]) -> str:
    start = _to_year_month(item)
    end = _end_year_month(item)
    if item.get("is_current") and not end:
        end = "至今"
    if start and end:
        return f"{start}–{end}"
    if start:
        return f"自 {start}"
    return end or "—"


def _truncate(text: Optional[str], max_len: int = 240) -> Optional[str]:
    if not text:
        return text
    text = str(text).strip()
    return text if len(text) <= max_len else text[: max_len - 1] + "…"


def _slugify(value: str) -> str:
    value = value.lower().strip()
    value = re.sub(r"[^a-z0-9\-_.]+", "-", value)
    value = re.sub(r"-+", "-", value).strip("-")
    return value or "profile"


def _extract_identifier_from_url(url: str) -> Optional[str]:
    m = re.search(r"linkedin\.com/in/([^/?#]+)/?", url)
    return m.group(1) if m else None


def pretty_format(profile: Dict[str, Any]) -> str:
    lines: List[str] = []
    name = profile.get("name") or "(未提供)"
    headline = profile.get("headline") or ""
    location = profile.get("location") or ""
    lines.append(f"姓名：{name}")
    if headline:
        lines.append(f"头衔：{headline}")
    if location:
        lines.append(f"地点：{location}")
    lines.append("")

    # 教育
    lines.append("教育背景：")
    educations = profile.get("educations") or []
    if not educations:
        lines.append("  - 无")
    else:
        for edu in educations:
            sy = edu.get("start_year") or ""
            ey = edu.get("end_year") or ""
            period = f"{sy}–{ey}".strip("–") if (sy or ey) else "—"
            school = edu.get("school") or ""
            degree = edu.get("degree") or ""
            field = edu.get("field_of_study") or ""
            main = " - ".join([x for x in [school, degree] if x]) or school
            tail = f"（{field}）" if field else ""
            lines.append(f"  - {period} {main}{tail}")
    lines.append("")

    # 经历
    lines.append("职业经历：")
    experiences = profile.get("experiences") or []
    if not experiences:
        lines.append("  - 无")
    else:
        for exp in experiences:
            period = _format_period({
                "start_date_year": _safe_int(exp.get("start_date", "")[:4]),
                "start_date_month": _safe_int(exp.get("start_date", "")[5:7]),
                "end_date_year": None,
                "end_date_month": None,
                "is_current": exp.get("is_current"),
            })
            company = exp.get("company") or ""
            title = exp.get("title") or ""
            loc = exp.get("location") or ""
            header = " - ".join([x for x in [company, title] if x])
            loc_suffix = f"（{loc}）" if loc else ""
            lines.append(f"  - {period} {header}{loc_suffix}")
            summary = _truncate(exp.get("summary"), 200)
            if summary:
                lines.append(f"      摘要：{summary}")

    return "\n".join(lines)


def _safe_int(value: Optional[str]) -> Optional[int]:
    try:
        if value is None or value == "":
            return None
        return int(value)
    except Exception:
        return None


def save_to_excel(profile: Dict[str, Any], url: str) -> str:
    exports_dir = os.path.join(os.path.dirname(__file__), "exports")
    os.makedirs(exports_dir, exist_ok=True)

    identifier = _extract_identifier_from_url(url) or _slugify(profile.get("name") or "profile")
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"{identifier}_{ts}.xlsx"
    path = os.path.abspath(os.path.join(exports_dir, filename))

    wb = Workbook()

    # Profile sheet
    ws_info = wb.active
    ws_info.title = "Profile"
    ws_info.append(["Field", "Value"])
    ws_info.append(["Name", profile.get("name")])
    ws_info.append(["Headline", profile.get("headline")])
    ws_info.append(["Location", profile.get("location")])

    # Experiences sheet
    ws_exp = wb.create_sheet("Experiences")
    ws_exp.append(["Company", "Title", "Start", "End", "Current", "Location", "Summary"])
    for exp in profile.get("experiences") or []:
        # 尝试从 title/summary 中推断起止时间字符串已在 profile 中，保持原样
        start_val = exp.get("start_date")
        end_val = exp.get("end_date") or ("至今" if exp.get("is_current") else None)
        ws_exp.append([
            exp.get("company"),
            exp.get("title"),
            start_val,
            end_val,
            bool(exp.get("is_current")),
            exp.get("location"),
            exp.get("summary"),
        ])

    # Educations sheet
    ws_edu = wb.create_sheet("Educations")
    ws_edu.append(["School", "Degree", "Field", "StartYear", "EndYear", "Description"])
    for edu in profile.get("educations") or []:
        ws_edu.append([
            edu.get("school"),
            edu.get("degree"),
            edu.get("field_of_study"),
            edu.get("start_year"),
            edu.get("end_year"),
            edu.get("description"),
        ])

    wb.save(path)
    return path


class PrettyResponse(BaseModel):
    name: Optional[str]
    headline: Optional[str]
    location: Optional[str]
    text: str
    excel_path: str


@app.post("/parse/pretty", response_model=PrettyResponse)
def parse_and_pretty(req: ParseRequest) -> Dict[str, Any]:
    if not is_linkedin_profile_url(req.url):
        raise HTTPException(status_code=400, detail="请输入合法的 LinkedIn 个人主页链接（形如 https://www.linkedin.com/in/<username>/ ）")

    provider = resolve_provider(req.provider)
    if provider == "serpapi":
        serp_key = os.getenv("SERPAPI_KEY")
        if not serp_key:
            raise HTTPException(status_code=500, detail="缺少 SERPAPI_KEY 环境变量")
        data = fetch_profile_from_serpapi(req.url, serp_key)
        parsed = project_profile_from_serpapi(data)
    else:
        api_token = os.getenv("CONTACTOUT_API_TOKEN")
        if not api_token:
            raise HTTPException(status_code=500, detail="缺少 CONTACTOUT_API_TOKEN 环境变量")
        data = fetch_profile_from_contactout(req.url, api_token)
        parsed = project_profile(data)
    text = pretty_format(parsed)
    excel_path = save_to_excel(parsed, req.url)
    return {
        "name": parsed.get("name"),
        "headline": parsed.get("headline"),
        "location": parsed.get("location"),
        "text": text,
        "excel_path": excel_path,
    }


# ========== 文本解析模式（无需第三方API） ==========

_YEAR_RANGE_RE = re.compile(
    r"(?P<y1>19\d{2}|20\d{2})(?:[年/\-.](?P<m1>\d{1,2}))?\s*(?:-|–|至|到|~|—)\s*(?P<y2>19\d{2}|20\d{2}|至今|现在|Present)(?:[年/\-.](?P<m2>\d{1,2}))?",
    re.IGNORECASE,
)


def _normalize_period_from_text(s: str) -> Dict[str, Optional[str]]:
    m = _YEAR_RANGE_RE.search(s)
    if not m:
        return {"start": None, "end": None, "is_current": None}
    y1 = m.group("y1")
    m1 = m.group("m1")
    y2 = m.group("y2")
    m2 = m.group("m2")
    start = f"{y1}{('-' + m1.zfill(2)) if m1 else ''}"
    if y2 and y2.lower() in {"至今", "现在", "present"}:
        end = None
        current = True
    else:
        end = f"{y2}{('-' + m2.zfill(2)) if m2 else ''}" if y2 else None
        current = False
    return {"start": start, "end": end, "is_current": current}


def _split_sections(text: str) -> Dict[str, List[str]]:
    lines = [l.strip() for l in text.splitlines() if l.strip()]
    sections = {"edu": [], "exp": []}
    current = None
    for line in lines:
        low = line.lower()
        if any(k in low for k in ["education", "教育", "教育背景"]):
            current = "edu"
            continue
        if any(k in low for k in ["experience", "工作经历", "經歷", "职业经历"]):
            current = "exp"
            continue
        if current:
            sections[current].append(line)
    return sections


def parse_profile_from_text(req: TextParseRequest) -> Dict[str, Any]:
    sections = _split_sections(req.text)
    educations: List[Dict[str, Any]] = []
    for line in sections["edu"]:
        period = _normalize_period_from_text(line)
        # 粗略提取学校与学位
        degree_tokens = [
            "phd", "doctor", "博士", "master", "硕士", "msc", "m.s.", "bachelor", "学士", "b.sc", "b.s.",
        ]
        degree = None
        for tok in degree_tokens:
            if tok in line.lower():
                degree = tok
                break
        # 学校名取掉日期后的剩余部分的前半段
        school = re.split(_YEAR_RANGE_RE, line)[0].strip()
        educations.append({
            "school": school or None,
            "degree": degree,
            "field_of_study": None,
            "start_year": (period["start"] or "")[:4] or None,
            "end_year": (period["end"] or "")[:4] or None,
            "description": None,
        })

    experiences: List[Dict[str, Any]] = []
    for line in sections["exp"]:
        period = _normalize_period_from_text(line)
        # 简单双分：公司 - 职位
        parts = re.split(r"\s+-\s+|，|,", line, maxsplit=1)
        company = parts[0].strip() if parts else None
        title = parts[1].strip() if len(parts) > 1 else None
        experiences.append({
            "company": company,
            "title": title,
            "start_date": period["start"],
            "end_date": period["end"],
            "is_current": period["is_current"],
            "location": None,
            "summary": None,
        })

    return {
        "name": req.name,
        "headline": req.headline,
        "location": req.location,
        "experiences": experiences,
        "educations": educations,
    }


# ========== 浏览器抓取模式（使用 li_at Cookie） ==========

def _ensure_playwright_available():
    if sync_playwright is None:
        raise HTTPException(status_code=500, detail="未安装 Playwright，请先安装依赖并下载浏览器：pip install playwright && playwright install chromium")


def _scrape_profile_text_with_cookies(url: str, li_at: str, jsessionid: Optional[str], headless: bool) -> str:
    _ensure_playwright_available()
    with cast(object, sync_playwright)() as p:  # type: ignore
        browser = p.chromium.launch(headless=headless)
        context = browser.new_context()
        # 注入 Cookie
        cookies = [
            {
                "name": "li_at",
                "value": li_at,
                "domain": ".linkedin.com",
                "path": "/",
                "httpOnly": True,
                "secure": True,
            }
        ]
        if jsessionid:
            cookies.append({
                "name": "JSESSIONID",
                "value": jsessionid,
                "domain": ".linkedin.com",
                "path": "/",
                "httpOnly": True,
                "secure": True,
            })
        context.add_cookies(cookies)

        page = context.new_page()
        page.goto(url, wait_until="networkidle")

        # 简单抽取：页面所有可见文本
        content_text = page.locator("body").inner_text()

        # 关闭
        context.close()
        browser.close()
        return content_text


@app.post("/parse/browser", response_model=PrettyResponse)
def parse_via_browser(req: BrowserParseRequest) -> Dict[str, Any]:
    if not is_linkedin_profile_url(req.url):
        raise HTTPException(status_code=400, detail="请输入合法的 LinkedIn 个人主页链接（形如 https://www.linkedin.com/in/<username>/ ）")
    if not req.li_at:
        raise HTTPException(status_code=400, detail="li_at 不能为空")

    text = _scrape_profile_text_with_cookies(req.url, req.li_at, req.jsessionid, bool(req.headless))

    # 复用文本解析模式
    parsed = parse_profile_from_text(TextParseRequest(text=text))
    pretty = pretty_format(parsed)
    excel_path = save_to_excel(parsed, req.url)
    return {
        "name": parsed.get("name"),
        "headline": parsed.get("headline"),
        "location": parsed.get("location"),
        "text": pretty,
        "excel_path": excel_path,
    }


@app.post("/parse/text", response_model=PrettyResponse)
def parse_text(req: TextParseRequest) -> Dict[str, Any]:
    if not req.text or not req.text.strip():
        raise HTTPException(status_code=400, detail="text 不能为空")
    parsed = parse_profile_from_text(req)
    text = pretty_format(parsed)
    # 以 name 或 'profile' 作为导出名
    tmp_url = f"https://linkedin.com/in/{_slugify(parsed.get('name') or 'profile')}"
    excel_path = save_to_excel(parsed, tmp_url)
    return {
        "name": parsed.get("name"),
        "headline": parsed.get("headline"),
        "location": parsed.get("location"),
        "text": text,
        "excel_path": excel_path,
    }


